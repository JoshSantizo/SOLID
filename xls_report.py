from report import Report
from employee import Employee
from openpyxl import Workbook


class XLSReport(Report):
    def __init__(self, name: str):
        super().__init__(name)

    def generate_report(self):
        wb = Workbook()
        sheet = wb.active
        row: int = 1

        for employee in self.employees:
            sheet.cell(row=row, column=1, value=employee.get_name())
            row += 1
            for fecha in employee.get_attendance():
                sheet.cell(row=row, column=1, value=fecha.current_date)
                sheet.cell(row=row, column=2, value=fecha.is_on_time())
                row += 1
        wb.save(self.name)

    def add_employee(self, employee: Employee):
        self.employees.append(employee)
